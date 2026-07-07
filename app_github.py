import streamlit as st
import pandas as pd
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuración de la página con estética corporativa y de lectura limpia
st.set_page_config(page_title="Portal de Modulaciones", layout="wide")

st.markdown("""
    <style>
    div.row-widget.stRadio > div {
        flex-direction: row;
        background-color: #f1f3f5;
        padding: 10px 15px;
        border-radius: 6px;
        border: 1px solid #ced4da;
    }
    /* Estilos extra para los KPIs */
    div[data-testid="metric-container"] {
        background-color: #ffffff;
        border: 1px solid #e0e0e0;
        padding: 15px;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    </style>
""", unsafe_allow_html=True)

# Inicialización de variables en caché para las bases externas y control de generación
if 'df_ventanas' not in st.session_state:
    st.session_state['df_ventanas'] = None
if 'df_ventas_ext' not in st.session_state:
    st.session_state['df_ventas_ext'] = None
if 'df_camiones_mesas' not in st.session_state:
    st.session_state['df_camiones_mesas'] = None
if 'reporte_hojas_cargadas' not in st.session_state:
    st.session_state['reporte_hojas_cargadas'] = {}
if 'autocargado' not in st.session_state:
    st.session_state['autocargado'] = False

# Variables para controlar cuándo mostrar los reportes
if 'mostrar_reporte_gen' not in st.session_state:
    st.session_state['mostrar_reporte_gen'] = False
if 'last_file_gen' not in st.session_state:
    st.session_state['last_file_gen'] = None
if 'mostrar_reporte_foc' not in st.session_state:
    st.session_state['mostrar_reporte_foc'] = False
if 'last_files_foc' not in st.session_state:
    st.session_state['last_files_foc'] = None

def cargar_datos(file):
    if file.name.endswith('.xlsx'):
        return pd.read_excel(file)
    else:
        return pd.read_csv(file, sep=None, engine='python')

def traducir_estado(val):
    if pd.isna(val): return "SIN REGISTRO"
    val_str = str(val).upper().strip()
    if 'CONCLUDED' in val_str:
        return 'Entregado'
    elif 'IN_TREATMENT' in val_str or 'RESCHEDULED' in val_str:
        return 'Rechazo'
    elif 'NOT_STARTED' in val_str or 'ON_THE_WAY' in val_str or 'DELIVERY_STARTED' in val_str:
        return 'No Iniciado'
    return val_str

def color_status(val):
    if pd.isna(val):
        return ''
    val_str = str(val).upper()
    if 'RECHAZO' in val_str:
        return 'background-color: #FF0000; color: black; font-weight: bold;'
    elif 'NO INICIADO' in val_str:
        return 'background-color: #FFFFFF; color: black;'
    elif 'ENTREGADO' in val_str:
        return 'background-color: #A9D08E; color: black;'
    return ''

def estandarizar_ventana(texto):
    if pd.isna(texto):
        return "SIN ASIGNAR"
    
    t = str(texto).upper().strip()
    if "ANTES DE" in t:
        num = ''.join(filter(str.isdigit, t))
        if num: return f"00:00 - {num.zfill(2)}:00"
            
    if "DESPUÉS DE" in t or "DESPUES DE" in t:
        num = ''.join(filter(str.isdigit, t))
        if num: return f"{num.zfill(2)}:00 - 23:59"
            
    t = t.replace(' AM', '').replace(' PM', '').replace('AM', '').replace('PM', '')
    return t

# -------------------------------------------------------------
# NUEVAS FUNCIONES PARA INDICADORES (KPIs)
# -------------------------------------------------------------
def evaluar_otd(arribo_str, ventana_str):
    """Evalúa si la hora de arribo está dentro de la ventana horaria asignada."""
    if pd.isna(arribo_str) or pd.isna(ventana_str) or arribo_str == "-" or "SIN" in str(ventana_str).upper() or "NO" in str(ventana_str).upper():
        return None
    try:
        arr_time = datetime.strptime(str(arribo_str).strip(), '%H:%M:%S').time()
        partes = str(ventana_str).split('-')
        if len(partes) == 2:
            inicio_str = partes[0].strip()
            fin_str = partes[1].strip()
            
            inicio = datetime.strptime(inicio_str, '%H:%M').time() if len(inicio_str.split(':')) == 2 else datetime.strptime(inicio_str, '%H:%M:%S').time()
            
            if fin_str == "23:59":
                fin = datetime.strptime("23:59:59", '%H:%M:%S').time()
            elif len(fin_str.split(':')) == 2:
                fin = datetime.strptime(fin_str, '%H:%M').time()
            else:
                fin = datetime.strptime(fin_str, '%H:%M:%S').time()
                
            return inicio <= arr_time <= fin
    except:
        pass
    return None

def mostrar_kpis(df):
    """Calcula y renderiza las tarjetas de Resumen Ejecutivo."""
    st.markdown("#### Resumen Ejecutivo (KPIs)")
    total_pdvs = len(df)
    if total_pdvs == 0:
        st.info("No hay datos para calcular métricas con los filtros actuales.")
        st.markdown("---")
        return

    # 1. Efectividad (Strike Rate)
    entregados = (df['Motivo'].str.upper() == 'ENTREGADO').sum()
    efectividad = (entregados / total_pdvs) * 100

    # 2. Volumen Rechazado (HL)
    hl_rechazados = 0
    if 'HL' in df.columns:
        df_rechazos = df[df['Motivo'].str.upper() == 'RECHAZO']
        hl_rechazados = pd.to_numeric(df_rechazos['HL'], errors='coerce').sum()

    # 3. Drop Time
    tiempos = pd.to_numeric(df['Tiempo_Entrega_Min'], errors='coerce')
    tiempos_validos = tiempos[tiempos > 0]
    drop_time = tiempos_validos.mean() if not tiempos_validos.empty else 0

    # 4. Cumplimiento Ventana (OTD)
    otd_results = df.apply(lambda x: evaluar_otd(x.get('Hora_Arribo'), x.get('Ventana_Horaria')), axis=1)
    valid_otd = otd_results.dropna()
    otd_pct = (valid_otd.sum() / len(valid_otd)) * 100 if not valid_otd.empty else 0
    base_otd = len(valid_otd)

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Efectividad (Strike Rate)", f"{efectividad:.1f}%", f"{entregados} de {total_pdvs} PDVs")
    col2.metric("Volumen Rechazado", f"{hl_rechazados:.2f} HL", "Merma Operativa", delta_color="inverse")
    col3.metric("Drop Time Promedio", f"{drop_time:.1f} min", "Tiempo en PDV")
    
    otd_label = f"{otd_pct:.1f}%" if base_otd > 0 else "N/A"
    col4.metric("Cumplimiento (OTD)", otd_label, f"Base: {base_otd} ventanas válidas")
    st.markdown("---")

# -------------------------------------------------------------

def generar_excel_colores(df_simp):
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"
    
    fill_entregado = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
    fill_rechazo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    fill_no_iniciado = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_header = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Arial", size=11, color="000000")
    font_rechazo = Font(name="Arial", size=11, bold=True, color="000000")
    
    align_center = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style='thin', color='D9D9D9')
    border_celda = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    ws.append(list(df_simp.columns))
    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        
    for r_idx, row in enumerate(df_simp.values, start=2):
        ws.append(list(row))
        status_val = str(row[3]).upper() if len(row) > 3 else ""
        
        for c_idx in range(1, len(row) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = font_body
            cell.alignment = align_center
            cell.border = border_celda
            
            if c_idx == 4:  
                if 'ENTREGADO' in status_val:
                    cell.fill = fill_entregado
                elif 'RECHAZO' in status_val:
                    cell.fill = fill_rechazo
                    cell.font = font_rechazo
                else:
                    cell.fill = fill_no_iniciado
                    
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def ejecutar_sincronizacion(silencioso=False):
    bitacora_hojas = {}
    
    try:
        sheet_id_vh = "1OYlT2SVGqxM-C6h27GSrBEcjBOyCixwP"
        url_vh = f"https://docs.google.com/spreadsheets/d/{sheet_id_vh}/export?format=xlsx"
        xls_vh = pd.ExcelFile(url_vh)
        df_vh_list = []
        
        if "VH FIJAS LPZ" in xls_vh.sheet_names:
            df_lpz = pd.read_excel(xls_vh, sheet_name="VH FIJAS LPZ")
            bitacora_hojas["VH FIJAS LPZ"] = f"{len(df_lpz)} registros"
            if len(df_lpz.columns) >= 10:
                df_temp = df_lpz.iloc[:, [1, 9]].copy()
                df_temp.columns = ['PDV_Drive', 'Ventana_Tratada']
                df_vh_list.append(df_temp)
                
        if "VHs FIJAS EA" in xls_vh.sheet_names:
            df_ea = pd.read_excel(xls_vh, sheet_name="VHs FIJAS EA")
            bitacora_hojas["VHs FIJAS EA"] = f"{len(df_ea)} registros"
            if len(df_ea.columns) >= 6:
                df_temp = df_ea.iloc[:, [0, 5]].copy()
                df_temp.columns = ['PDV_Drive', 'Ventana_Tratada']
                df_vh_list.append(df_temp)
        
        if "VH FIJA VESPERTINAS LPZ" in xls_vh.sheet_names:
            df_vesp = pd.read_excel(xls_vh, sheet_name="VH FIJA VESPERTINAS LPZ")
            bitacora_hojas["VH FIJA VESPERTINAS LPZ"] = f"{len(df_vesp)} registros"
            if len(df_vesp.columns) >= 2:
                df_temp = df_vesp.iloc[:, [0, 1]].copy()
                df_temp.columns = ['PDV_Drive', 'Ventana_Tratada']
                df_vh_list.append(df_temp)
                
        if df_vh_list:
            df_consolidado_vh = pd.concat(df_vh_list, ignore_index=True)
            df_consolidado_vh.dropna(subset=['PDV_Drive'], inplace=True)
            df_consolidado_vh['PDV_Drive'] = df_consolidado_vh['PDV_Drive'].astype(str).str.strip().str.replace('.0', '', regex=False)
            df_consolidado_vh['Ventana_Tratada'] = df_consolidado_vh['Ventana_Tratada'].apply(estandarizar_ventana)
            df_consolidado_vh = df_consolidado_vh.drop_duplicates(subset=['PDV_Drive'], keep='first')
            st.session_state['df_ventanas'] = df_consolidado_vh
    except Exception as e:
        if not silencioso: st.error(f"Error en sincronización de Ventanas: {e}")

    try:
        sheet_id_ventas = "1zIllojDvh23QUOP8afJbxD66I5Ly6tgY"
        url_ventas = f"https://docs.google.com/spreadsheets/d/{sheet_id_ventas}/export?format=xlsx"
        xls_ventas = pd.ExcelFile(url_ventas)
        
        if "Clientes" in xls_ventas.sheet_names and "datos_ventas" in xls_ventas.sheet_names:
            df_clientes_ext = pd.read_excel(xls_ventas, sheet_name="Clientes")
            bitacora_hojas["Clientes"] = f"{len(df_clientes_ext)} registros"
            df_clientes_ext = df_clientes_ext.iloc[:, [1, 8]].copy()
            df_clientes_ext.columns = ['PDV_Drive_Ventas', 'Territorio']
            df_clientes_ext['PDV_Drive_Ventas'] = df_clientes_ext['PDV_Drive_Ventas'].astype(str).str.strip().str.replace('.0', '', regex=False)
            df_clientes_ext['Territorio'] = df_clientes_ext['Territorio'].astype(str).str.strip()
            
            df_datos_ventas = pd.read_excel(xls_ventas, sheet_name="datos_ventas")
            bitacora_hojas["datos_ventas"] = f"{len(df_datos_ventas)} registros"
            nombres_columnas_ventas = df_datos_ventas.columns
            df_datos_ventas = df_datos_ventas.iloc[:, [0, 1, 2, 3, 4]].copy()
            df_datos_ventas.columns = ['Territorio', nombres_columnas_ventas[1], nombres_columnas_ventas[2], nombres_columnas_ventas[3], nombres_columnas_ventas[4]]
            df_datos_ventas['Territorio'] = df_datos_ventas['Territorio'].astype(str).str.strip()
            
            df_ventas_consolidadas = pd.merge(df_clientes_ext, df_datos_ventas, on='Territorio', how='left')
            df_ventas_consolidadas = df_ventas_consolidadas.drop_duplicates(subset=['PDV_Drive_Ventas'], keep='first')
            st.session_state['df_ventas_ext'] = df_ventas_consolidadas
            
        if "estructura_camiones" in xls_ventas.sheet_names:
            df_mesas_raw = pd.read_excel(xls_ventas, sheet_name="estructura_camiones")
            bitacora_hojas["estructura_camiones"] = "Procesada"
            
            mesas_map = {
                0: "Power", 1: "Alfas", 2: "Los Pioneros", 
                3: "Guerreros", 4: "Turbos", 5: "Gladiadores", 6: "UDC"
            }
            camiones_list = []
            
            for col_idx in range(min(7, len(df_mesas_raw.columns))):
                nombre_mesa = mesas_map[col_idx]
                col_data = df_mesas_raw.iloc[:, col_idx].dropna().astype(str).str.strip().str.upper()
                for camion_cod in col_data:
                    if camion_cod and camion_cod != "NAN":
                        camiones_list.append({"Camion_Ref": camion_cod, "Mesa": nombre_mesa})
            
            if camiones_list:
                df_camiones_mesas = pd.DataFrame(camiones_list).drop_duplicates(subset=["Camion_Ref"])
                st.session_state['df_camiones_mesas'] = df_camiones_mesas

        st.session_state['reporte_hojas_cargadas'] = bitacora_hojas
        if not silencioso: st.success("Bases de datos externas sincronizadas correctamente.")
    except Exception as e:
        if not silencioso: st.error(f"Error en sincronización de Ventas o Mesas: {e}")

if not st.session_state['autocargado']:
    with st.spinner("Inicializando portal y sincronizando bases de datos corporativas..."):
        ejecutar_sincronizacion(silencioso=True)
        st.session_state['autocargado'] = True

col_encabezado, col_control_sync = st.columns([3, 1])
with col_encabezado:
    st.title("Portal de Modulaciones y Seguimiento de Entregas")
    st.markdown("Herramienta automatizada para el cruce de pedidos, tiempos, ventanas horarias, mesas y métricas de venta.")

with col_control_sync:
    st.markdown("<div style='padding-top: 25px;'></div>", unsafe_allow_html=True)
    if st.button("Sincronizar Drive Manual", type="primary", use_container_width=True, key="sync_top_button"):
        with st.spinner("Actualizando datos operativos..."):
            ejecutar_sincronizacion(silencioso=False)

st.markdown("---")
tab_general, tab_focus = st.tabs(["Modulaciones General", "Modulaciones Focus"])

# ==========================================
# PESTAÑA 1: MODULACIONES GENERAL
# ==========================================
with tab_general:
    st.subheader("Base Despachos (Vista General)")
    file_entregas_gen = st.file_uploader("Cargar archivo de Despachos", type=["xlsx", "csv"], key="entregas_general")

    if file_entregas_gen:
        try:
            if st.session_state['last_file_gen'] != file_entregas_gen.name:
                st.session_state['mostrar_reporte_gen'] = False
                st.session_state['last_file_gen'] = file_entregas_gen.name

            df_origen_gen = cargar_datos(file_entregas_gen)
            df_resultado_gen = df_origen_gen.copy()
            cols_2_gen = df_resultado_gen.columns.tolist()
            
            col_cruce_f_gen = next((c for c in cols_2_gen if 'poc_exter' in str(c).lower()), cols_2_gen[5] if len(cols_2_gen) > 5 else cols_2_gen[0])
            col_datos_d_gen = next((c for c in cols_2_gen if 'driver' in str(c).lower()), cols_2_gen[3] if len(cols_2_gen) > 3 else cols_2_gen[0])
            col_status_i_gen = next((c for c in cols_2_gen if 'status' in str(c).lower()), cols_2_gen[8] if len(cols_2_gen) > 8 else cols_2_gen[0])
            
            # Ajuste de Columnas Solicitadas: Detalle Rechazo (W=22) y HL (AO=40)
            col_motivo_x_gen = cols_2_gen[22] if len(cols_2_gen) > 22 else cols_2_gen[-1]
            col_hl_gen = cols_2_gen[40] if len(cols_2_gen) > 40 else cols_2_gen[-1]
            
            col_arrived_gen = next((c for c in cols_2_gen if 'arrived' in str(c).lower()), cols_2_gen[20] if len(cols_2_gen) > 20 else cols_2_gen[-1])
            col_finished_gen = next((c for c in cols_2_gen if 'finished' in str(c).lower()), cols_2_gen[21] if len(cols_2_gen) > 21 else cols_2_gen[-1])

            with st.expander("Verificar Parámetros de Extracción (General)", expanded=True):
                c1_g, c2_g, c3_g, c4_g = st.columns(4)
                with c1_g: col_cruce_2_gen_sel = st.selectbox("Identificador PDV:", cols_2_gen, index=cols_2_gen.index(col_cruce_f_gen), key="sel_pdv_gen")
                with c2_g: col_mostrar_d_gen_sel = st.selectbox("Conductor:", cols_2_gen, index=cols_2_gen.index(col_datos_d_gen), key="sel_drv_gen")
                with c3_g: col_arrived_sel_gen_sel = st.selectbox("Llegada:", cols_2_gen, index=cols_2_gen.index(col_arrived_gen), key="sel_arr_gen")
                with c4_g: col_finished_sel_gen_sel = st.selectbox("Salida:", cols_2_gen, index=cols_2_gen.index(col_finished_gen), key="sel_fin_gen")

            if st.button("Generar Reporte General", type="primary", use_container_width=True, key="btn_gen_reporte_general"):
                st.session_state['mostrar_reporte_gen'] = True

            if st.session_state.get('mostrar_reporte_gen', False):
                df_resultado_gen['_llave_cruce_'] = df_resultado_gen[col_cruce_2_gen_sel].astype(str).str.strip().str.replace('.0', '', regex=False)

                if st.session_state['df_ventanas'] is not None:
                    df_vh_gen = st.session_state['df_ventanas'].copy()
                    df_resultado_gen = pd.merge(df_resultado_gen, df_vh_gen, left_on='_llave_cruce_', right_on='PDV_Drive', how='left')
                    col_ventana_view_gen = 'Ventana_Tratada'
                    df_resultado_gen[col_ventana_view_gen] = df_resultado_gen[col_ventana_view_gen].fillna("SIN ASIGNAR")
                else:
                    df_resultado_gen['Ventana_Tratada'] = "NO CARGADA"
                    col_ventana_view_gen = 'Ventana_Tratada'

                columnas_ventas_agregadas_gen = []
                if st.session_state['df_ventas_ext'] is not None:
                    df_ve_gen = st.session_state['df_ventas_ext'].copy()
                    df_resultado_gen = pd.merge(df_resultado_gen, df_ve_gen, left_on='_llave_cruce_', right_on='PDV_Drive_Ventas', how='left')
                    df_resultado_gen['Territorio'] = df_resultado_gen['Territorio'].fillna("SIN TERRITORIO")
                    columnas_ventas_agregadas_gen = [c for c in df_ve_gen.columns if c not in ['PDV_Drive_Ventas', 'Territorio']]
                    for c in columnas_ventas_agregadas_gen:
                        df_resultado_gen[c] = df_resultado_gen[c].fillna("-")
                else:
                    df_resultado_gen['Territorio'] = "NO CARGADO"

                if col_mostrar_d_gen_sel in df_resultado_gen.columns:
                    df_resultado_gen[col_mostrar_d_gen_sel] = df_resultado_gen[col_mostrar_d_gen_sel].fillna("SIN DATOS")
                    split_data_gen = df_resultado_gen[col_mostrar_d_gen_sel].astype(str).str.split('-', expand=True)
                    if split_data_gen.shape[1] > 1:
                        df_resultado_gen['Camion'] = split_data_gen[1].str.strip().str.upper()
                    else:
                        df_resultado_gen['Camion'] = split_data_gen[0].str.strip().str.upper()
                else:
                    df_resultado_gen['Camion'] = "NO ENCONTRADO"

                if st.session_state['df_camiones_mesas'] is not None:
                    df_mesas_gen = st.session_state['df_camiones_mesas'].copy()
                    df_resultado_gen = pd.merge(df_resultado_gen, df_mesas_gen, left_on='Camion', right_on='Camion_Ref', how='left')
                    df_resultado_gen['Mesa'] = df_resultado_gen['Mesa'].fillna("SIN MESA")
                else:
                    df_resultado_gen['Mesa'] = "NO CARGADA"

                df_resultado_gen[col_arrived_sel_gen_sel] = pd.to_datetime(df_resultado_gen[col_arrived_sel_gen_sel], errors='coerce')
                df_resultado_gen[col_finished_sel_gen_sel] = pd.to_datetime(df_resultado_gen[col_finished_sel_gen_sel], errors='coerce')
                df_resultado_gen['Hora_Arribo'] = df_resultado_gen[col_arrived_sel_gen_sel].dt.strftime('%H:%M:%S').fillna("-")
                df_resultado_gen['Tiempo_Entrega_Min'] = (df_resultado_gen[col_finished_sel_gen_sel] - df_resultado_gen[col_arrived_sel_gen_sel]).dt.total_seconds() / 60
                df_resultado_gen['Tiempo_Entrega_Min'] = df_resultado_gen['Tiempo_Entrega_Min'].round(2).fillna("-")

                df_resultado_gen[col_status_i_gen] = df_resultado_gen[col_status_i_gen].apply(traducir_estado)

                columnas_base_vista_gen = [col_cruce_2_gen_sel, 'Camion', 'Mesa', 'Territorio', col_status_i_gen, col_ventana_view_gen, 'Hora_Arribo', 'Tiempo_Entrega_Min', col_motivo_x_gen, col_hl_gen]
                columnas_totales_vista_gen = columnas_base_vista_gen + columnas_ventas_agregadas_gen
                
                vista_final_gen = df_resultado_gen[columnas_totales_vista_gen].rename(columns={
                    col_cruce_2_gen_sel: 'PDV',
                    col_status_i_gen: 'Motivo',
                    col_ventana_view_gen: 'Ventana_Horaria',
                    col_motivo_x_gen: 'Detalle_Rechazo',
                    col_hl_gen: 'HL'
                })
                
                condicion_motivo_gen = vista_final_gen['Motivo'].str.upper() == 'RECHAZO'
                vista_final_gen['Detalle_Rechazo'] = vista_final_gen['Detalle_Rechazo'].where(condicion_motivo_gen, "-").fillna("SIN DETALLE")

                st.markdown("### Controles de Búsqueda y Filtrado")
                col_search_g, col_filter_g = st.columns([1, 2])
                with col_search_g:
                    search_pdv_g = st.text_input("Buscar por Código PDV:", key="src_gen")
                with col_filter_g:
                    status_filter_g = st.radio(
                        "Segmentación de Estado:", 
                        ["Todos", "Entregado", "Rechazo", "No Iniciado"], 
                        horizontal=True,
                        key="rad_gen"
                    )
                
                opcion_tabla_gen = st.radio("Formato de tabla en pantalla:", ["Reporte Completo", "Formato Simplificado (4 Columnas)"], horizontal=True, key="op_tbl_gen")

                if search_pdv_g:
                    vista_final_gen = vista_final_gen[vista_final_gen['PDV'].astype(str).str.contains(search_pdv_g, case=False, na=False)]
                    
                if status_filter_g != "Todos":
                    vista_final_gen = vista_final_gen[vista_final_gen['Motivo'] == status_filter_g]

                columnas_simplificadas_gen = ['PDV', 'Camion', 'Territorio', 'Motivo']
                tabla_a_mostrar_gen = vista_final_gen[columnas_simplificadas_gen] if opcion_tabla_gen == "Formato Simplificado (4 Columnas)" else vista_final_gen

                # MOSTRAR KPIs GENERAL
                mostrar_kpis(vista_final_gen)

                st.markdown("### Consolidado General")
                if hasattr(tabla_a_mostrar_gen.style, 'map'):
                    styled_df_gen = tabla_a_mostrar_gen.style.map(color_status, subset=['Motivo'])
                else:
                    styled_df_gen = tabla_a_mostrar_gen.style.applymap(color_status, subset=['Motivo'])
                    
                st.dataframe(styled_df_gen, use_container_width=True)
                
                st.markdown("---")
                st.markdown("### Cumplimiento Ventanas Horarias")
                if st.session_state['reporte_hojas_cargadas']:
                    st.info(f"Estado de bases externas en Drive: {str(st.session_state['reporte_hojas_cargadas']).replace('{','').replace('}','')}")
                else:
                    st.warning("Estructura de Ventanas Horarias: NO CARGADA desde la nube.")
                
                st.markdown("---")
                st.markdown("### Exportación de Reportes (General)")
                
                col_dl1_gen, col_dl2_gen = st.columns(2)
                
                with col_dl1_gen:
                    csv_data_gen = vista_final_gen.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        label="Descargar Reporte Completo (CSV)",
                        data=csv_data_gen,
                        file_name="reporte_general_completo.csv",
                        mime="text/csv",
                        type="primary",
                        key="btn_dl_gen_comp"
                    )
                    
                with col_dl2_gen:
                    buffer_gen = generar_excel_colores(vista_final_gen[columnas_simplificadas_gen])
                    st.download_button(
                        label="Descargar Formato Excel (Con colores)",
                        data=buffer_gen,
                        file_name="reporte_general_simplificado.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="btn_dl_gen_simp"
                    )

        except Exception as e:
            st.error(f"Se presentó un error en el procesamiento general: {e}")

# ==========================================
# PESTAÑA 2: MODULACIONES FOCUS
# ==========================================
with tab_focus:
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("1. Base Pedidos")
        file_clientes = st.file_uploader("Cargar archivo de Pedidos (Excel/CSV)", type=["xlsx", "csv"], key="clientes")
        tratamiento_clientes = st.radio(
            "Formato de Base Pedidos:", 
            ["Normal", "Requiere Tratamiento"], 
            horizontal=True,
            key="tratamiento_pedidos_focus"
        )

    with col2:
        st.subheader("2. Base Despachos")
        file_entregas = st.file_uploader("Cargar archivo de Despachos (Excel/CSV)", type=["xlsx", "csv"], key="entregas")

    if file_clientes and file_entregas:
        try:
            current_foc_names = file_clientes.name + "_" + file_entregas.name
            if st.session_state['last_files_foc'] != current_foc_names:
                st.session_state['mostrar_reporte_foc'] = False
                st.session_state['last_files_foc'] = current_foc_names

            df_clientes_raw = cargar_datos(file_clientes)
            df_entregas_raw = cargar_datos(file_entregas)

            df_clientes = df_clientes_raw.copy()
            df_entregas = df_entregas_raw.copy()

            if tratamiento_clientes == "Requiere Tratamiento":
                if len(df_clientes.columns) == 1:
                    col_name = df_clientes.columns[0]
                    header_parts = col_name.split(',')
                    df_clientes = df_clientes[col_name].astype(str).str.split(',', expand=True)
                    if len(header_parts) == df_clientes.shape[1]:
                        df_clientes.columns = header_parts
                
                col_0 = df_clientes.columns[0]
                split_guion = df_clientes[col_0].astype(str).str.split('-', expand=True)
                
                df_tratado = pd.DataFrame()
                df_tratado['COD'] = split_guion[0].str.strip() if split_guion.shape[1] > 0 else ""
                df_tratado['NOM'] = split_guion[1].str.strip() if split_guion.shape[1] > 1 else ""
                df_tratado['PDV'] = split_guion[2].str.strip() if split_guion.shape[1] > 2 else ""
                
                for c in df_clientes.columns[1:]:
                    df_tratado[c] = df_clientes[c]
                    
                df_clientes = df_tratado
                cols_1 = df_clientes.columns.tolist()
                col_pdv = 'PDV' 
            else:
                cols_1 = df_clientes.columns.tolist()
                col_pdv = next((c for c in cols_1 if str(c).upper() == 'PDV'), cols_1[1] if len(cols_1) > 1 else cols_1[0])

            cols_2 = df_entregas.columns.tolist()
            col_cruce_f = next((c for c in cols_2 if 'poc_exter' in str(c).lower()), cols_2[5] if len(cols_2) > 5 else cols_2[0])
            col_datos_d = next((c for c in cols_2 if 'driver' in str(c).lower()), cols_2[3] if len(cols_2) > 3 else cols_2[0])
            col_status_i = next((c for c in cols_2 if 'status' in str(c).lower()), cols_2[8] if len(cols_2) > 8 else cols_2[0])
            
            # Ajuste de Columnas Solicitadas: Detalle Rechazo (W=22) y HL (AO=40)
            col_motivo_x = cols_2[22] if len(cols_2) > 22 else cols_2[-1]
            col_hl = cols_2[40] if len(cols_2) > 40 else cols_2[-1]
            
            col_arrived = next((c for c in cols_2 if 'arrived' in str(c).lower()), cols_2[20] if len(cols_2) > 20 else cols_2[-1])
            col_finished = next((c for c in cols_2 if 'finished' in str(c).lower()), cols_2[21] if len(cols_2) > 21 else cols_2[-1])

            with st.expander("Verificar Parámetros de Cruce (Focus)", expanded=True):
                st.markdown("**Bases 1 y 2 (Pedidos y Despachos)**")
                c1_f, c2_f, c3_f, c4_f, c5_f = st.columns(5)
                with c1_f: col_cruce_1 = st.selectbox("Identificador (Base 1):", cols_1, index=cols_1.index(col_pdv), key="sel_foc_1")
                with c2_f: col_cruce_2 = st.selectbox("Buscar en (Base 2):", cols_2, index=cols_2.index(col_cruce_f), key="sel_foc_2")
                with c3_f: col_mostrar_d = st.selectbox("Conductor:", cols_2, index=cols_2.index(col_datos_d), key="sel_foc_3")
                with c4_f: col_arrived_sel = st.selectbox("Llegada:", cols_2, index=cols_2.index(col_arrived), key="sel_foc_4")
                with c5_f: col_finished_sel = st.selectbox("Salida:", cols_2, index=cols_2.index(col_finished), key="sel_foc_5")

            if st.button("Generar Reporte Focus", type="primary", use_container_width=True, key="btn_gen_reporte_focus"):
                st.session_state['mostrar_reporte_foc'] = True

            if st.session_state.get('mostrar_reporte_foc', False):
                df_clientes['_llave_cruce_'] = df_clientes[col_cruce_1].astype(str).str.strip().str.replace('.0', '', regex=False)
                df_entregas['_llave_cruce_'] = df_entregas[col_cruce_2].astype(str).str.strip().str.replace('.0', '', regex=False)

                df_entregas_subset = df_entregas[[col_cruce_2, '_llave_cruce_', col_mostrar_d, col_status_i, col_motivo_x, col_arrived_sel, col_finished_sel, col_hl]].drop_duplicates(subset=['_llave_cruce_'])
                df_resultado = pd.merge(
                    df_clientes[[col_cruce_1, '_llave_cruce_']], 
                    df_entregas_subset, 
                    on='_llave_cruce_', 
                    how='left'
                )

                if st.session_state['df_ventanas'] is not None:
                    df_vh = st.session_state['df_ventanas'].copy()
                    df_resultado = pd.merge(df_resultado, df_vh, left_on='_llave_cruce_', right_on='PDV_Drive', how='left')
                    col_ventana_view = 'Ventana_Tratada'
                    df_resultado[col_ventana_view] = df_resultado[col_ventana_view].fillna("SIN ASIGNAR")
                else:
                    df_resultado['Ventana_Tratada'] = "NO CARGADA"
                    col_ventana_view = 'Ventana_Tratada'

                columnas_ventas_agregadas = []
                if st.session_state['df_ventas_ext'] is not None:
                    df_ve = st.session_state['df_ventas_ext'].copy()
                    df_resultado = pd.merge(df_resultado, df_ve, left_on='_llave_cruce_', right_on='PDV_Drive_Ventas', how='left')
                    df_resultado['Territorio'] = df_resultado['Territorio'].fillna("SIN TERRITORIO")
                    columnas_ventas_agregadas = [c for c in df_ve.columns if c not in ['PDV_Drive_Ventas', 'Territorio']]
                    for c in columnas_ventas_agregadas:
                        df_resultado[c] = df_resultado[c].fillna("-")
                else:
                    df_resultado['Territorio'] = "NO CARGADO"
                
                if col_mostrar_d in df_resultado.columns:
                    df_resultado[col_mostrar_d] = df_resultado[col_mostrar_d].fillna("SIN DATOS")
                    split_data = df_resultado[col_mostrar_d].astype(str).str.split('-', expand=True)
                    if split_data.shape[1] > 1:
                        df_resultado['Camion'] = split_data[1].str.strip().str.upper()
                    else:
                        df_resultado['Camion'] = split_data[0].str.strip().str.upper()
                else:
                    df_resultado['Camion'] = "NO ENCONTRADO"

                if st.session_state['df_camiones_mesas'] is not None:
                    df_mesas_foc = st.session_state['df_camiones_mesas'].copy()
                    df_resultado = pd.merge(df_resultado, df_mesas_foc, left_on='Camion', right_on='Camion_Ref', how='left')
                    df_resultado['Mesa'] = df_resultado['Mesa'].fillna("SIN MESA")
                else:
                    df_resultado['Mesa'] = "NO CARGADA"

                df_resultado[col_arrived_sel] = pd.to_datetime(df_resultado[col_arrived_sel], errors='coerce')
                df_resultado[col_finished_sel] = pd.to_datetime(df_resultado[col_finished_sel], errors='coerce')
                df_resultado['Hora_Arribo'] = df_resultado[col_arrived_sel].dt.strftime('%H:%M:%S').fillna("-")
                df_resultado['Tiempo_Entrega_Min'] = (df_resultado[col_finished_sel] - df_resultado[col_arrived_sel]).dt.total_seconds() / 60
                df_resultado['Tiempo_Entrega_Min'] = df_resultado['Tiempo_Entrega_Min'].round(2).fillna("-")

                df_resultado[col_status_i] = df_resultado[col_status_i].apply(traducir_estado)

                columnas_base_vista = [col_cruce_1, 'Camion', 'Mesa', 'Territorio', col_status_i, col_ventana_view, 'Hora_Arribo', 'Tiempo_Entrega_Min', col_motivo_x, col_hl]
                columnas_totales_vista = columnas_base_vista + columnas_ventas_agregadas
                
                vista_final = df_resultado[columnas_totales_vista].rename(columns={
                    col_cruce_1: 'PDV',
                    col_status_i: 'Motivo',
                    col_ventana_view: 'Ventana_Horaria',
                    col_motivo_x: 'Detalle_Rechazo',
                    col_hl: 'HL'
                })
                
                condicion_motivo = vista_final['Motivo'].str.upper() == 'RECHAZO'
                vista_final['Detalle_Rechazo'] = vista_final['Detalle_Rechazo'].where(condicion_motivo, "-").fillna("SIN DETALLE")

                st.markdown("### Controles de Búsqueda y Filtrado")
                col_search_f, col_filter_f = st.columns([1, 2])
                with col_search_f:
                    search_pdv = st.text_input("Buscar por Código PDV:", key="src_foc")
                with col_filter_f:
                    status_filter = st.radio(
                        "Segmentación de Estado:", 
                        ["Todos", "Entregado", "Rechazo", "No Iniciado"], 
                        horizontal=True,
                        key="rad_foc"
                    )
                
                opcion_tabla_foc = st.radio("Formato de tabla en pantalla:", ["Reporte Completo", "Formato Simplificado (4 Columnas)"], horizontal=True, key="op_tbl_foc")

                if search_pdv:
                    vista_final = vista_final[vista_final['PDV'].astype(str).str.contains(search_pdv, case=False, na=False)]
                    
                if status_filter != "Todos":
                    vista_final = vista_final[vista_final['Motivo'] == status_filter]

                columnas_simplificadas_foc = ['PDV', 'Camion', 'Territorio', 'Motivo']
                tabla_a_mostrar_foc = vista_final[columnas_simplificadas_foc] if opcion_tabla_foc == "Formato Simplificado (4 Columnas)" else vista_final

                # MOSTRAR KPIs FOCUS
                mostrar_kpis(vista_final)

                st.markdown("### Consolidado Focus")
                if hasattr(tabla_a_mostrar_foc.style, 'map'):
                    styled_df = tabla_a_mostrar_foc.style.map(color_status, subset=['Motivo'])
                else:
                    styled_df = tabla_a_mostrar_foc.style.applymap(color_status, subset=['Motivo'])
                    
                st.dataframe(styled_df, use_container_width=True)
                
                st.markdown("---")
                st.markdown("### Cumplimiento Ventanas Horarias")
                if st.session_state['reporte_hojas_cargadas']:
                    st.info(f"Estado de bases externas en Drive: {str(st.session_state['reporte_hojas_cargadas']).replace('{','').replace('}','')}")
                else:
                    st.warning("Estructura de Ventanas Horarias: NO CARGADA desde la nube.")
                
                st.markdown("---")
                st.markdown("### Exportación de Reportes (Focus)")
                
                col_dl1_foc, col_dl2_foc = st.columns(2)
                
                with col_dl1_foc:
                    csv_data_foc = vista_final.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        label="Descargar Reporte Completo (CSV)",
                        data=csv_data_foc,
                        file_name="reporte_focus_completo.csv",
                        mime="text/csv",
                        type="primary",
                        key="btn_dl_foc_comp"
                    )
                    
                with col_dl2_foc:
                    buffer_foc = generar_excel_colores(vista_final[columnas_simplificadas_foc])
                    st.download_button(
                        label="Descargar Formato Excel (Con colores)",
                        data=buffer_foc,
                        file_name="reporte_focus_simplificado.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="btn_dl_foc_simp"
                    )

        except Exception as e:
            st.error(f"Se presentó un error en el procesamiento focus: {e}")
